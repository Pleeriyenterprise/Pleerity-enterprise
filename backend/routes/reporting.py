"""
Full Reporting System - Export & Scheduling
Provides PDF, CSV, Excel exports and scheduled report delivery
"""
import io
import csv
import json
import uuid
import os
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any, Literal
from fastapi import APIRouter, HTTPException, Depends, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr

from middleware import admin_route_guard
from database import database
from models.core import AuditAction, UserRole
from utils.audit import create_audit_log
import logging

# Excel support
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF support
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/admin/reports", tags=["Reports"])


# ============================================
# Report Types
# ============================================

class ReportType(str):
    REVENUE = "revenue"
    ORDERS = "orders"
    CLIENTS = "clients"
    LEADS = "leads"
    COMPLIANCE = "compliance"
    SLA = "sla"
    ENABLEMENT = "enablement"
    CONSENT = "consent"


class ExportFormat(str):
    CSV = "csv"
    JSON = "json"
    PDF = "pdf"
    XLSX = "xlsx"


# ============================================
# Request/Response Models
# ============================================

class ReportRequest(BaseModel):
    report_type: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    period: str = "30d"
    format: str = "csv"
    filters: Dict[str, Any] = {}


class ScheduledReportCreate(BaseModel):
    name: str = Field(..., max_length=100)
    report_type: str
    frequency: Literal["daily", "weekly", "monthly"]
    recipients: List[EmailStr]
    format: str = "csv"
    filters: Dict[str, Any] = {}
    enabled: bool = True


class ScheduledReportResponse(BaseModel):
    schedule_id: str
    name: str
    report_type: str
    frequency: str
    recipients: List[str]
    format: str
    filters: Dict[str, Any]
    enabled: bool
    last_run: Optional[datetime]
    next_run: Optional[datetime]
    created_at: datetime
    created_by: str


# ============================================
# Helper Functions
# ============================================

def generate_schedule_id() -> str:
    return f"SCHED-{uuid.uuid4().hex[:12].upper()}"


def generate_report_id() -> str:
    return f"RPT-{uuid.uuid4().hex[:12].upper()}"


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def parse_date(date_str: str) -> datetime:
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except ValueError:
        return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)


def get_date_range(period: str) -> tuple:
    """Get start and end dates for a period."""
    end = now_utc()
    
    if period == "today":
        start = end.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "7d":
        start = end - timedelta(days=7)
    elif period == "30d":
        start = end - timedelta(days=30)
    elif period == "90d":
        start = end - timedelta(days=90)
    elif period == "ytd":
        start = end.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = end - timedelta(days=30)
    
    return start, end


# ============================================
# Data Fetchers
# ============================================

async def fetch_revenue_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch revenue report data."""
    db = database.get_db()
    
    query = {
        "created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()},
        "stripe_payment_status": "paid"
    }
    
    if filters.get("service_code"):
        query["service_code"] = filters["service_code"]
    
    cursor = db.orders.find(query, {"_id": 0}).sort("created_at", -1)
    orders = await cursor.to_list(10000)
    
    data = []
    for order in orders:
        data.append({
            "Order ID": order.get("order_id"),
            "Date": order.get("created_at", "")[:10],
            "Client ID": order.get("client_id"),
            "Service": order.get("service_name", order.get("service_code")),
            "Amount (£)": (order.get("pricing", {}).get("total_pence", 0) or 0) / 100,
            "Status": order.get("status"),
            "Payment Status": order.get("stripe_payment_status"),
        })
    
    return data


async def fetch_orders_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch orders report data."""
    db = database.get_db()
    
    query = {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}
    
    if filters.get("status"):
        query["status"] = filters["status"]
    
    cursor = db.orders.find(query, {"_id": 0}).sort("created_at", -1)
    orders = await cursor.to_list(10000)
    
    data = []
    for order in orders:
        data.append({
            "Order ID": order.get("order_id"),
            "Date": order.get("created_at", "")[:10],
            "Client ID": order.get("client_id"),
            "Service": order.get("service_name"),
            "Status": order.get("status"),
            "Amount (£)": (order.get("pricing", {}).get("total_pence", 0) or 0) / 100,
            "Fast Track": "Yes" if order.get("fast_track") else "No",
            "Printed Copy": "Yes" if order.get("printed_copy") else "No",
        })
    
    return data


async def fetch_clients_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch clients report data."""
    db = database.get_db()
    
    query = {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}
    
    cursor = db.clients.find(query, {"_id": 0}).sort("created_at", -1)
    clients = await cursor.to_list(10000)
    
    data = []
    for client in clients:
        data.append({
            "Client ID": client.get("client_id"),
            "Name": client.get("name"),
            "Email": client.get("email"),
            "Phone": client.get("phone"),
            "Created": client.get("created_at", "")[:10],
            "Plan": client.get("plan_code"),
            "Status": client.get("onboarding_status"),
            "Properties": client.get("property_count", 0),
        })
    
    return data


async def fetch_leads_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch leads report data."""
    db = database.get_db()
    
    query = {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}
    
    if filters.get("stage"):
        query["stage"] = filters["stage"]
    if filters.get("intent_score"):
        query["intent_score"] = {"$gte": filters["intent_score"]}
    
    cursor = db.leads.find(query, {"_id": 0}).sort("created_at", -1)
    leads = await cursor.to_list(10000)
    
    data = []
    for lead in leads:
        data.append({
            "Lead ID": lead.get("lead_id"),
            "Email": lead.get("email"),
            "Name": lead.get("name"),
            "Created": lead.get("created_at", "")[:10],
            "Source": lead.get("source_platform"),
            "Stage": lead.get("stage"),
            "Intent Score": lead.get("intent_score"),
            "Converted": "Yes" if lead.get("converted_to_client") else "No",
        })
    
    return data


async def fetch_compliance_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch compliance report data."""
    db = database.get_db()
    
    # Get all properties
    cursor = db.properties.find({}, {"_id": 0})
    properties = await cursor.to_list(10000)
    
    data = []
    for prop in properties:
        data.append({
            "Property ID": prop.get("property_id"),
            "Address": prop.get("address", {}).get("line1", ""),
            "Postcode": prop.get("address", {}).get("postcode", ""),
            "Client ID": prop.get("client_id"),
            "Compliance %": prop.get("compliance_score", 0),
            "Status": prop.get("compliance_status", "UNKNOWN"),
            "Total Requirements": prop.get("total_requirements", 0),
            "Met Requirements": prop.get("met_requirements", 0),
        })
    
    return data


async def fetch_enablement_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch enablement report data."""
    db = database.get_db()
    
    query = {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}
    
    if filters.get("status"):
        query["status"] = filters["status"]
    if filters.get("category"):
        query["category"] = filters["category"]
    
    cursor = db.enablement_actions.find(query, {"_id": 0}).sort("created_at", -1)
    actions = await cursor.to_list(10000)
    
    data = []
    for action in actions:
        data.append({
            "Action ID": action.get("action_id"),
            "Date": action.get("created_at", "")[:19],
            "Client ID": action.get("client_id"),
            "Event Type": action.get("event_type"),
            "Category": action.get("category"),
            "Channel": action.get("channel"),
            "Status": action.get("status"),
            "Title": action.get("rendered_title", "")[:50],
        })
    
    return data


async def fetch_consent_data(start: datetime, end: datetime, filters: dict) -> List[dict]:
    """Fetch consent report data."""
    db = database.get_db()
    
    query = {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}
    
    cursor = db.consent_events.find(query, {"_id": 0}).sort("created_at", -1)
    events = await cursor.to_list(10000)
    
    data = []
    for event in events:
        prefs = event.get("preferences", {})
        data.append({
            "Event ID": event.get("event_id"),
            "Date": event.get("created_at", "")[:19],
            "Session ID": event.get("session_id"),
            "Action": event.get("event_type"),
            "Analytics": "Yes" if prefs.get("analytics") else "No",
            "Marketing": "Yes" if prefs.get("marketing") else "No",
            "Functional": "Yes" if prefs.get("functional") else "No",
        })
    
    return data


# Data fetcher mapping
DATA_FETCHERS = {
    "revenue": fetch_revenue_data,
    "orders": fetch_orders_data,
    "clients": fetch_clients_data,
    "leads": fetch_leads_data,
    "compliance": fetch_compliance_data,
    "enablement": fetch_enablement_data,
    "consent": fetch_consent_data,
}


# ============================================
# Export Formatters
# ============================================

def format_csv(data: List[dict]) -> io.StringIO:
    """Format data as CSV."""
    output = io.StringIO()
    
    if not data:
        output.write("No data available\n")
        return output
    
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    output.seek(0)
    return output


def format_json(data: List[dict]) -> io.StringIO:
    """Format data as JSON."""
    output = io.StringIO()
    json.dump({"data": data, "count": len(data), "generated_at": now_utc().isoformat()}, output, indent=2, default=str)
    output.seek(0)
    return output


def format_xlsx(data: List[dict], report_type: str, start: datetime, end: datetime) -> io.BytesIO:
    """Format data as Excel XLSX."""
    output = io.BytesIO()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{report_type.title()} Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add date range row
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Period: {start.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Add generated timestamp
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated: {now_utc().strftime('%d %b %Y %H:%M:%S')} UTC"
    ws['A3'].font = Font(size=9, color="666666")
    ws['A3'].alignment = Alignment(horizontal="center")
    
    if not data:
        ws['A5'] = "No data available for this period"
        wb.save(output)
        output.seek(0)
        return output
    
    # Write headers (row 5)
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row_data in enumerate(data, 6):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border
            # Format currency cells
            if "£" in str(header) or "Amount" in str(header):
                cell.number_format = '£#,##0.00'
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    wb.save(output)
    output.seek(0)
    return output


def format_pdf(data: List[dict], report_type: str, start: datetime, end: datetime) -> io.BytesIO:
    """Format data as PDF."""
    output = io.BytesIO()
    
    # Use landscape for more columns
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"{report_type.title()} Report", title_style))
    
    # Date range
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=5,
        alignment=1
    )
    elements.append(Paragraph(
        f"Period: {start.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}",
        subtitle_style
    ))
    elements.append(Paragraph(
        f"Generated: {now_utc().strftime('%d %b %Y %H:%M:%S')} UTC",
        subtitle_style
    ))
    elements.append(Spacer(1, 20))
    
    if not data:
        elements.append(Paragraph("No data available for this period", styles['Normal']))
        doc.build(elements)
        output.seek(0)
        return output
    
    # Prepare table data
    headers = list(data[0].keys())
    table_data = [headers]
    
    # Limit columns for PDF readability (max 8 columns)
    if len(headers) > 8:
        headers = headers[:8]
        table_data = [headers]
    
    for row in data[:500]:  # Limit rows for PDF
        row_values = [str(row.get(h, ""))[:30] for h in headers]  # Truncate long values
        table_data.append(row_values)
    
    # Calculate column widths
    available_width = landscape(A4)[0] - 60  # Total page width minus margins
    col_width = available_width / len(headers)
    
    # Create table
    table = Table(table_data, colWidths=[col_width] * len(headers))
    
    # Style the table
    table_style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(table_style)
    elements.append(table)
    
    # Add row count note if truncated
    if len(data) > 500:
        elements.append(Spacer(1, 10))
        note_style = ParagraphStyle(
            'Note',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey
        )
        elements.append(Paragraph(
            f"Note: Showing first 500 of {len(data)} rows. Export as CSV or Excel for full data.",
            note_style
        ))
    
    doc.build(elements)
    output.seek(0)
    return output


# ============================================
# API Endpoints
# ============================================

@router.get("/types")
async def get_report_types(admin: dict = Depends(admin_route_guard)):
    """Get available report types."""
    return {
        "types": [
            {"value": "revenue", "label": "Revenue Report", "description": "Revenue by orders and services"},
            {"value": "orders", "label": "Orders Report", "description": "All orders with status"},
            {"value": "clients", "label": "Clients Report", "description": "Client listing and details"},
            {"value": "leads", "label": "Leads Report", "description": "Lead pipeline data"},
            {"value": "compliance", "label": "Compliance Report", "description": "Property compliance status"},
            {"value": "enablement", "label": "Enablement Report", "description": "Customer enablement actions"},
            {"value": "consent", "label": "Consent Report", "description": "Cookie consent events"},
        ],
        "formats": [
            {"value": "csv", "label": "CSV", "description": "Spreadsheet compatible"},
            {"value": "xlsx", "label": "Excel", "description": "Microsoft Excel format"},
            {"value": "pdf", "label": "PDF", "description": "Professional document"},
            {"value": "json", "label": "JSON", "description": "Developer friendly"},
        ],
        "periods": [
            {"value": "today", "label": "Today"},
            {"value": "7d", "label": "Last 7 Days"},
            {"value": "30d", "label": "Last 30 Days"},
            {"value": "90d", "label": "Last 90 Days"},
            {"value": "ytd", "label": "Year to Date"},
            {"value": "custom", "label": "Custom Range"},
        ]
    }


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    admin: dict = Depends(admin_route_guard)
):
    """Generate a report and return as download."""
    
    if request.report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {request.report_type}")
    
    # Determine date range
    if request.start_date and request.end_date:
        start = parse_date(request.start_date)
        end = parse_date(request.end_date)
    else:
        start, end = get_date_range(request.period)
    
    # Fetch data
    fetcher = DATA_FETCHERS[request.report_type]
    data = await fetcher(start, end, request.filters)
    
    # Format output based on requested format
    if request.format == "xlsx":
        output = format_xlsx(data, request.report_type, start, end)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"
        content = output.getvalue()
    elif request.format == "pdf":
        output = format_pdf(data, request.report_type, start, end)
        media_type = "application/pdf"
        extension = "pdf"
        content = output.getvalue()
    elif request.format == "json":
        output = format_json(data)
        media_type = "application/json"
        extension = "json"
        content = output.getvalue().encode('utf-8') if isinstance(output.getvalue(), str) else output.getvalue()
    else:  # Default to CSV
        output = format_csv(data)
        media_type = "text/csv"
        extension = "csv"
        content = output.getvalue().encode('utf-8') if isinstance(output.getvalue(), str) else output.getvalue()
    
    filename = f"{request.report_type}_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.{extension}"
    
    # Audit log
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report",
        resource_id=generate_report_id(),
        metadata={
            "report_type": request.report_type,
            "format": request.format,
            "period": f"{start.isoformat()} to {end.isoformat()}",
            "row_count": len(data)
        }
    )
    
    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/preview/{report_type}")
async def preview_report(
    report_type: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    period: str = "30d",
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(admin_route_guard)
):
    """Preview report data without downloading."""
    
    if report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")
    
    # Determine date range
    if start_date and end_date:
        start = parse_date(start_date)
        end = parse_date(end_date)
    else:
        start, end = get_date_range(period)
    
    # Fetch data
    fetcher = DATA_FETCHERS[report_type]
    data = await fetcher(start, end, {})
    
    return {
        "report_type": report_type,
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "total_rows": len(data),
        "preview": data[:limit],
        "columns": list(data[0].keys()) if data else []
    }


# ============================================
# Scheduled Reports
# ============================================

@router.get("/schedules")
async def list_scheduled_reports(admin: dict = Depends(admin_route_guard)):
    """List all scheduled reports."""
    db = database.get_db()
    
    cursor = db.report_schedules.find({}, {"_id": 0}).sort("created_at", -1)
    schedules = await cursor.to_list(100)
    
    return {"schedules": schedules, "total": len(schedules)}


@router.post("/schedules")
async def create_scheduled_report(
    request: ScheduledReportCreate,
    admin: dict = Depends(admin_route_guard)
):
    """Create a new scheduled report."""
    db = database.get_db()
    
    if request.report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {request.report_type}")
    
    schedule_id = generate_schedule_id()
    now = now_utc()
    
    # Calculate next run based on frequency
    if request.frequency == "daily":
        next_run = now.replace(hour=8, minute=0, second=0, microsecond=0) + timedelta(days=1)
    elif request.frequency == "weekly":
        days_until_monday = (7 - now.weekday()) % 7
        next_run = now.replace(hour=8, minute=0, second=0, microsecond=0) + timedelta(days=days_until_monday or 7)
    else:  # monthly
        if now.month == 12:
            next_run = now.replace(year=now.year + 1, month=1, day=1, hour=8, minute=0, second=0, microsecond=0)
        else:
            next_run = now.replace(month=now.month + 1, day=1, hour=8, minute=0, second=0, microsecond=0)
    
    schedule = {
        "schedule_id": schedule_id,
        "name": request.name,
        "report_type": request.report_type,
        "frequency": request.frequency,
        "recipients": request.recipients,
        "format": request.format,
        "filters": request.filters,
        "enabled": request.enabled,
        "last_run": None,
        "next_run": next_run.isoformat(),
        "created_at": now.isoformat(),
        "created_by": admin.get("portal_user_id"),
    }
    
    await db.report_schedules.insert_one(schedule)
    
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report_schedule",
        resource_id=schedule_id,
        metadata={"action": "create", "name": request.name, "frequency": request.frequency}
    )
    
    if "_id" in schedule:
        del schedule["_id"]
    return schedule


@router.put("/schedules/{schedule_id}/toggle")
async def toggle_scheduled_report(
    schedule_id: str,
    admin: dict = Depends(admin_route_guard)
):
    """Enable or disable a scheduled report."""
    db = database.get_db()
    
    schedule = await db.report_schedules.find_one({"schedule_id": schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    new_status = not schedule.get("enabled", True)
    
    await db.report_schedules.update_one(
        {"schedule_id": schedule_id},
        {"$set": {"enabled": new_status}}
    )
    
    return {"schedule_id": schedule_id, "enabled": new_status}


@router.delete("/schedules/{schedule_id}")
async def delete_scheduled_report(
    schedule_id: str,
    admin: dict = Depends(admin_route_guard)
):
    """Delete a scheduled report."""
    db = database.get_db()
    
    result = await db.report_schedules.delete_one({"schedule_id": schedule_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report_schedule",
        resource_id=schedule_id,
        metadata={"action": "delete"}
    )
    
    return {"success": True}


# ============================================
# Report History
# ============================================

@router.get("/history")
async def get_report_history(
    limit: int = Query(50, ge=1, le=200),
    admin: dict = Depends(admin_route_guard)
):
    """Get report generation history from audit logs."""
    db = database.get_db()
    
    cursor = db.audit_logs.find(
        {"resource_type": "report"},
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit)
    
    logs = await cursor.to_list(limit)
    
    return {"history": logs, "total": len(logs)}


# ============================================
# Email Delivery
# ============================================

async def send_report_email(
    recipients: List[str],
    report_type: str,
    format: str,
    data: List[dict],
    start: datetime,
    end: datetime,
    schedule_name: str = None
) -> dict:
    """Generate report and send via NotificationOrchestrator with attachment."""
    from services.notification_orchestrator import notification_orchestrator
    import base64
    # Generate file attachment
    if format == "xlsx":
        output = format_xlsx(data, report_type, start, end)
        file_content = output.getvalue()
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"
    elif format == "pdf":
        output = format_pdf(data, report_type, start, end)
        file_content = output.getvalue()
        content_type = "application/pdf"
        extension = "pdf"
    else:  # Default CSV
        output = format_csv(data)
        file_content = output.getvalue().encode('utf-8')
        content_type = "text/csv"
        extension = "csv"
    
    filename = f"{report_type}_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.{extension}"
    
    # Build email body
    period_str = f"{start.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}"
    subject = f"{'[' + schedule_name + '] ' if schedule_name else ''}{report_type.title()} Report - {period_str}"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333;">
        <h2 style="color: #1E3A5F;">Your Report is Ready</h2>
        <p>Please find attached your <strong>{report_type.title()} Report</strong>.</p>
        <table style="border-collapse: collapse; margin: 20px 0;">
            <tr>
                <td style="padding: 8px 16px; border: 1px solid #ddd; background: #f5f5f5;"><strong>Report Type</strong></td>
                <td style="padding: 8px 16px; border: 1px solid #ddd;">{report_type.title()}</td>
            </tr>
            <tr>
                <td style="padding: 8px 16px; border: 1px solid #ddd; background: #f5f5f5;"><strong>Period</strong></td>
                <td style="padding: 8px 16px; border: 1px solid #ddd;">{period_str}</td>
            </tr>
            <tr>
                <td style="padding: 8px 16px; border: 1px solid #ddd; background: #f5f5f5;"><strong>Total Records</strong></td>
                <td style="padding: 8px 16px; border: 1px solid #ddd;">{len(data):,}</td>
            </tr>
            <tr>
                <td style="padding: 8px 16px; border: 1px solid #ddd; background: #f5f5f5;"><strong>Format</strong></td>
                <td style="padding: 8px 16px; border: 1px solid #ddd;">{format.upper()}</td>
            </tr>
            <tr>
                <td style="padding: 8px 16px; border: 1px solid #ddd; background: #f5f5f5;"><strong>Generated</strong></td>
                <td style="padding: 8px 16px; border: 1px solid #ddd;">{now_utc().strftime('%d %b %Y %H:%M:%S')} UTC</td>
            </tr>
        </table>
        <p style="color: #666; font-size: 12px;">
            This is an automated report from Pleerity Enterprise. 
            If you no longer wish to receive these reports, please update your schedule settings.
        </p>
    </body>
    </html>
    """
    
    text_body = f"""
Your Report is Ready

Report Type: {report_type.title()}
Period: {period_str}
Total Records: {len(data):,}
Format: {format.upper()}
Generated: {now_utc().strftime('%d %b %Y %H:%M:%S')} UTC

Please find the report attached to this email.

This is an automated report from Pleerity Enterprise.
    """
    
    results = []
    attachments = [{"Name": filename, "Content": base64.b64encode(file_content).decode("utf-8"), "ContentType": content_type}]
    for recipient in recipients:
        try:
            idempotency_key = f"SCHEDULED_REPORT_{report_type}_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}_{recipient}"
            result = await notification_orchestrator.send(
                template_key="SCHEDULED_REPORT",
                client_id=None,
                context={
                    "recipient": recipient,
                    "subject": subject,
                    "message": html_body,
                    "attachments": attachments,
                },
                idempotency_key=idempotency_key,
                event_type="report_email",
            )
            if result.outcome in ("sent", "duplicate_ignored"):
                results.append({"recipient": recipient, "status": "sent", "message_id": result.message_id})
            else:
                results.append({"recipient": recipient, "status": "failed", "error": result.error_message or result.block_reason})
        except Exception as e:
            logger.error(f"Failed to send report email to {recipient}: {e}")
            results.append({"recipient": recipient, "status": "failed", "error": str(e)})
    return {"results": results, "filename": filename}


@router.post("/schedules/{schedule_id}/run")
async def run_scheduled_report_now(
    schedule_id: str,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(admin_route_guard)
):
    """Manually trigger a scheduled report to run immediately."""
    db = database.get_db()
    
    schedule = await db.report_schedules.find_one({"schedule_id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    report_type = schedule["report_type"]
    if report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")
    
    # Get data for last period based on frequency
    now = now_utc()
    if schedule["frequency"] == "daily":
        start = now - timedelta(days=1)
    elif schedule["frequency"] == "weekly":
        start = now - timedelta(days=7)
    else:  # monthly
        start = now - timedelta(days=30)
    end = now
    
    # Fetch data
    fetcher = DATA_FETCHERS[report_type]
    data = await fetcher(start, end, schedule.get("filters", {}))
    
    # Send email
    email_result = await send_report_email(
        recipients=schedule["recipients"],
        report_type=report_type,
        format=schedule.get("format", "csv"),
        data=data,
        start=start,
        end=end,
        schedule_name=schedule["name"]
    )
    
    # Update last run time
    await db.report_schedules.update_one(
        {"schedule_id": schedule_id},
        {"$set": {"last_run": now.isoformat()}}
    )
    
    # Log the execution
    await db.report_executions.insert_one({
        "execution_id": generate_report_id(),
        "schedule_id": schedule_id,
        "schedule_name": schedule["name"],
        "report_type": report_type,
        "format": schedule.get("format", "csv"),
        "recipients": schedule["recipients"],
        "row_count": len(data),
        "triggered_by": admin.get("portal_user_id"),
        "trigger_type": "manual",
        "email_results": email_result["results"],
        "executed_at": now.isoformat()
    })
    
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report_schedule",
        resource_id=schedule_id,
        metadata={"action": "manual_run", "recipients": schedule["recipients"], "row_count": len(data)}
    )
    
    return {
        "success": True,
        "schedule_id": schedule_id,
        "report_type": report_type,
        "recipients": schedule["recipients"],
        "row_count": len(data),
        "email_results": email_result["results"]
    }


@router.get("/executions")
async def get_report_executions(
    schedule_id: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    admin: dict = Depends(admin_route_guard)
):
    """Get scheduled report execution history."""
    db = database.get_db()
    
    query = {}
    if schedule_id:
        query["schedule_id"] = schedule_id
    
    cursor = db.report_executions.find(query, {"_id": 0}).sort("executed_at", -1).limit(limit)
    executions = await cursor.to_list(limit)
    
    return {"executions": executions, "total": len(executions)}


# ============================================
# Report Sharing - Public Links
# ============================================

class ShareReportRequest(BaseModel):
    """Create a shareable report link"""
    report_type: str
    format: str = "pdf"
    period: str = "30d"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    expires_in_days: int = Field(default=7, ge=1, le=30)
    name: Optional[str] = None  # Optional name for the share


class ShareLinkResponse(BaseModel):
    share_id: str
    share_url: str
    report_type: str
    format: str
    expires_at: str
    created_at: str


@router.post("/share", response_model=ShareLinkResponse)
async def create_share_link(
    request: ShareReportRequest,
    admin: dict = Depends(admin_route_guard)
):
    """Create a time-limited shareable link for a report."""
    db = database.get_db()
    
    if request.report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {request.report_type}")
    
    # Determine date range
    if request.start_date and request.end_date:
        start = parse_date(request.start_date)
        end = parse_date(request.end_date)
    else:
        start, end = get_date_range(request.period)
    
    share_id = f"SHR-{uuid.uuid4().hex[:16].upper()}"
    now = now_utc()
    expires_at = now + timedelta(days=request.expires_in_days)
    
    share_record = {
        "share_id": share_id,
        "report_type": request.report_type,
        "format": request.format,
        "period": request.period,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "name": request.name or f"{request.report_type.title()} Report",
        "created_by": admin.get("portal_user_id"),
        "created_at": now.isoformat(),
        "expires_at": expires_at.isoformat(),
        "access_count": 0,
        "is_active": True
    }
    
    await db.report_shares.insert_one(share_record)
    
    # Build public URL (using frontend route)
    base_url = os.environ.get("PUBLIC_URL", "")
    share_url = f"{base_url}/shared/report/{share_id}"
    
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report_share",
        resource_id=share_id,
        metadata={"action": "create_share", "report_type": request.report_type, "expires_in_days": request.expires_in_days}
    )
    
    return ShareLinkResponse(
        share_id=share_id,
        share_url=share_url,
        report_type=request.report_type,
        format=request.format,
        expires_at=expires_at.isoformat(),
        created_at=now.isoformat()
    )


@router.get("/shares")
async def list_share_links(
    active_only: bool = True,
    limit: int = Query(50, ge=1, le=200),
    admin: dict = Depends(admin_route_guard)
):
    """List all shareable report links."""
    db = database.get_db()
    
    query = {}
    if active_only:
        query["is_active"] = True
        query["expires_at"] = {"$gt": now_utc().isoformat()}
    
    cursor = db.report_shares.find(query, {"_id": 0}).sort("created_at", -1).limit(limit)
    shares = await cursor.to_list(limit)
    
    return {"shares": shares, "total": len(shares)}


@router.delete("/shares/{share_id}")
async def revoke_share_link(
    share_id: str,
    admin: dict = Depends(admin_route_guard)
):
    """Revoke a shareable report link."""
    db = database.get_db()
    
    result = await db.report_shares.update_one(
        {"share_id": share_id},
        {"$set": {"is_active": False, "revoked_at": now_utc().isoformat(), "revoked_by": admin.get("portal_user_id")}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Share link not found")
    
    await create_audit_log(
        action=AuditAction.ADMIN_ACTION,
        actor_role=UserRole.ROLE_ADMIN,
        actor_id=admin.get("portal_user_id"),
        resource_type="report_share",
        resource_id=share_id,
        metadata={"action": "revoke_share"}
    )
    
    return {"success": True}


# ============================================
# Public Report Access (No Auth Required)
# ============================================

public_router = APIRouter(prefix="/api/public/reports", tags=["Public Reports"])


@public_router.get("/shared/{share_id}")
async def get_shared_report_info(share_id: str):
    """Get information about a shared report (public)."""
    db = database.get_db()
    
    share = await db.report_shares.find_one({"share_id": share_id}, {"_id": 0})
    
    if not share:
        raise HTTPException(status_code=404, detail="Share link not found")
    
    if not share.get("is_active", False):
        raise HTTPException(status_code=410, detail="This share link has been revoked")
    
    expires_at = datetime.fromisoformat(share["expires_at"].replace("Z", "+00:00"))
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="This share link has expired")
    
    return {
        "share_id": share_id,
        "name": share.get("name"),
        "report_type": share["report_type"],
        "format": share["format"],
        "period": f"{share['start_date'][:10]} to {share['end_date'][:10]}",
        "expires_at": share["expires_at"]
    }


@public_router.get("/shared/{share_id}/download")
async def download_shared_report(share_id: str):
    """Download a shared report (public)."""
    db = database.get_db()
    
    share = await db.report_shares.find_one({"share_id": share_id})
    
    if not share:
        raise HTTPException(status_code=404, detail="Share link not found")
    
    if not share.get("is_active", False):
        raise HTTPException(status_code=410, detail="This share link has been revoked")
    
    expires_at = datetime.fromisoformat(share["expires_at"].replace("Z", "+00:00"))
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="This share link has expired")
    
    report_type = share["report_type"]
    if report_type not in DATA_FETCHERS:
        raise HTTPException(status_code=400, detail="Invalid report type")
    
    # Parse dates
    start = datetime.fromisoformat(share["start_date"].replace("Z", "+00:00"))
    end = datetime.fromisoformat(share["end_date"].replace("Z", "+00:00"))
    
    # Fetch data
    fetcher = DATA_FETCHERS[report_type]
    data = await fetcher(start, end, {})
    
    # Format output
    format_type = share.get("format", "pdf")
    
    if format_type == "xlsx":
        output = format_xlsx(data, report_type, start, end)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"
        content = output.getvalue()
    elif format_type == "pdf":
        output = format_pdf(data, report_type, start, end)
        media_type = "application/pdf"
        extension = "pdf"
        content = output.getvalue()
    elif format_type == "json":
        output = format_json(data)
        media_type = "application/json"
        extension = "json"
        content = output.getvalue().encode('utf-8')
    else:  # csv
        output = format_csv(data)
        media_type = "text/csv"
        extension = "csv"
        content = output.getvalue().encode('utf-8')
    
    filename = f"{report_type}_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.{extension}"
    
    # Update access count
    await db.report_shares.update_one(
        {"share_id": share_id},
        {"$inc": {"access_count": 1}, "$set": {"last_accessed": now_utc().isoformat()}}
    )
    
    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
